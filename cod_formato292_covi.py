# -*- coding: utf-8 -*-
"""Cod_formato292_COVI.ipynb

Automatically generated by Colab.

Original file is located at
    https://colab.research.google.com/drive/1jB5W934N1CX7maD2ZDu9gTT2e8Gus_4J

# Cargar librerias
"""

import pandas as pd
import numpy as np
import openpyxl
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timedelta

"""# Cargar bases de datos"""

# Configuración de pandas para mostrar más dígitos en la salida
pd.set_option('display.float_format', '{:.22f}'.format)

# Imprime el directorio actual
print("Directorio inicial:", os.getcwd())

# Cambiar al directorio 'FORMATO_290'
os.chdir('..')

# Confirma el cambio de directorio
print("Directorio después de moverse hacia arriba:", os.getcwd())

ruta = r"D:\OneDrive - Allianz\6_Procesos TRANSMISIONES\Fuentes de Información COVI.xlsx"

BALANCE_WP = pd.read_excel(ruta,sheet_name="BALANCE_WP",usecols="A:E",skiprows=5,dtype=str)

# Renombrar columnas para que coincidan con la estructura anterior
BALANCE_WP.columns = ["CUENTA", "NOMBRE", "VALOR", "VALOR2", "VALOR MONEDA EXTRANJERA"]

# Cargamos el workbook y seleccionamos la hoja
wb = openpyxl.load_workbook('Bases\Plantilla_Capital_Minimo_COVI.xlsx')
ws = wb['Hoja1']

"""## CAPITAL MINIMO ACREDITADO UNIDAD DE CAPTURA 1"""

# Función para escribir valores en las celdas
def escribir_valor_excel(fila, columna, valor):
    ws.cell(row=fila, column=columna).value = valor if valor != 0 else 0

# Cálculo de los valores de cada cuenta y escritura en Excel
c_310500 = BALANCE_WP[(BALANCE_WP['CUENTA'] == "310500")].groupby(['CUENTA'])['VALOR'].sum().reset_index()
d18 = float(c_310500['VALOR'].iloc[0]) if not c_310500.empty else 0
escribir_valor_excel(18, 4, d18)  # Escribir en fila 18, columna 4

c_311500 = BALANCE_WP[(BALANCE_WP['CUENTA'] == "311500")].groupby(['CUENTA'])['VALOR'].sum().reset_index()
d19 = float(c_311500['VALOR'].iloc[0]) if not c_311500.empty else 0
escribir_valor_excel(19, 4, d19)  # Escribir en fila 19, columna 4

c_315500 = BALANCE_WP[(BALANCE_WP['CUENTA'] == "315500")].groupby(['CUENTA'])['VALOR'].sum().reset_index()
d20 = float(c_315500['VALOR'].iloc[0]) if not c_315500.empty else 0
escribir_valor_excel(20, 4, d20)  # Escribir en fila 20, columna 4

c_320000 = BALANCE_WP[(BALANCE_WP['CUENTA'] == "320000")].groupby(['CUENTA'])['VALOR'].sum().reset_index()
d21 = float(c_320000['VALOR'].iloc[0]) if not c_320000.empty else 0
escribir_valor_excel(21, 4, d21)  # Escribir en fila 21, columna 4

c_380500 = BALANCE_WP[(BALANCE_WP['CUENTA'] == "380500")].groupby(['CUENTA'])['VALOR'].sum().reset_index()
d22 = float(c_380500['VALOR'].iloc[0]) if not c_380500.empty else 0
escribir_valor_excel(22, 4, d22)  # Escribir en fila 22, columna 4

c_390500 = BALANCE_WP[(BALANCE_WP['CUENTA'] == "390500")].groupby(['CUENTA'])['VALOR'].sum().reset_index()
d23 = float(c_390500['VALOR'].iloc[0]) if not c_390500.empty else 0
escribir_valor_excel(23, 4, d23)  # Escribir en fila 23, columna 4

c_391000 = BALANCE_WP[(BALANCE_WP['CUENTA'] == "391000")].groupby(['CUENTA'])['VALOR'].sum().reset_index()
d24 = float(c_391000['VALOR'].iloc[0]) if not c_391000.empty else 0
escribir_valor_excel(24, 4, d24)  # Escribir en fila 24, columna 4

c_392000 = BALANCE_WP[(BALANCE_WP['CUENTA'] == "392000")].groupby(['CUENTA'])['VALOR'].sum().reset_index()
d25 = float(c_392000['VALOR'].iloc[0]) if not c_392000.empty else 0
escribir_valor_excel(25, 4, d25)  # Escribir en fila 25, columna 4

# Calcular el total d27_999 y escribir en la celda correspondiente
d26_999 = (d18 + d19 + d20 + d21 + d22 + d23 - d24 - d25)

# Escribir el total en fila 26, columna 4
escribir_valor_excel(26, 4, d26_999)

"""## CAPITAL MÍNIMO  DE FUNCIONAMIENTO REQUERIDO UNIDAD DE CAPTURA 2"""

# Obtener los valores de las celdas D30, D31, D32, D33 (column 4, rows 30 to 33)
d29 = ws.cell(row=29, column=4).value or 0
d30 = ws.cell(row=30, column=4).value or 0
d31 = ws.cell(row=31, column=4).value or 0
d32 = ws.cell(row=32, column=4).value or 0

# Realizar la suma de los valores de D30 a D33
d33 = d29 + d30 + d31 + d32

# Escribir el resultado de la suma en la celda D34 (fila 34, columna 4)
ws.cell(row=33, column=4).value = d33

"""## EXCESO O DEFECTO EN EL CAPITAL MINIMO REQUERIDO UNIDAD DE CAPTURA 3

"""

d26 = ws.cell(row=26, column=4).value or 0  # Si la celda está vacía, asignar 0
d33 = ws.cell(row=33, column=4).value or 0  # Si la celda está vacía, asignar 0

# Realizar la operación D27 - D34
d36 = d26 - d33

# Escribir el resultado en la celda D36
ws.cell(row=36, column=4).value = d36

"""## RESULTADOS DEL EJERCICIO	UNIDAD DE CAPTURA 4

"""

c_391500 = BALANCE_WP[(BALANCE_WP['CUENTA'] == "391500")].groupby(['CUENTA'])['VALOR'].sum().reset_index()
d39 = float(c_391500['VALOR'].iloc[0]) if not c_391500.empty else 0
ws.cell(row=39, column=4).value = d39

wb.save('Bases/Capital Minimo COVI.xlsx')
wb.close()

"""## PROFORMA TXT"""

# Cargar el archivo Excel
excel_file = "D:/OneDrive - Allianz/6_Procesos TRANSMISIONES/FORMATO 292_CAPITAL MINIMO/Bases/Capital Minimo COVI.xlsx"
df = pd.read_excel(excel_file, header=14)

# Lista de subcuentas válidas
subcuentas_validas = ['005', '010', '012', '015', '020', '035', '999', '032', '034']

# Función para generar el consecutivo de la línea
def generar_consecutivo(num):
    return str(num).zfill(5)

# Función para agregar el signo al valor y corregir el manejo de negativos
def formatear_valor(valor):
    try:
        # Si el valor es NaN o no es convertible, tratarlo como 0
        valor = float(valor) if pd.notna(valor) else 0
    except ValueError:
        # Si no se puede convertir, tratarlo como 0
        valor = 0

    # Revisar si el valor es positivo o negativo
    signo = '+' if valor >= 0 else '-'
    valor_abs = abs(valor)
    # Formatear el valor correctamente con el signo adecuado
    valor_formateado = f"{signo}{str(int(valor_abs)).zfill(20)}"  # Convertir a entero para zfill
    return valor_formateado

# Inicializar la lista para almacenar las líneas
lineas = []

# Obtener la fecha correcta (mes vencido)
fecha_actual = datetime.now()
fecha_mes_vencido = fecha_actual.replace(day=1) - timedelta(days=1)
fecha_str = fecha_mes_vencido.strftime("%d%m%Y")  # Formato día-mes-año

# Encabezados fijos de la primera y segunda línea
registro_inicial = f"00001114000001{fecha_str}00019SVIDCOLSECS0137"
segunda_linea = "000022000001100000"

# Agregar las primeras dos líneas a la lista
lineas.append(registro_inicial)
lineas.append(segunda_linea)

# Filtrar solo las filas que tienen subcuentas y unidades de captura válidas
df_filtrado = df[df['SUBCUENTA'].isin(subcuentas_validas)]

# Recorrer cada fila del DataFrame filtrado
numero_registro = 2  # Iniciar en 2 porque las dos primeras líneas están fijas

for index, row in df_filtrado.iterrows():
    subcuenta = str(row['SUBCUENTA']).zfill(3)  # Subcuenta debe tener 3 caracteres
    unidad_captura = str(int(row['UNIDAD DE CAPTURA'])).zfill(2)  # Unidad de captura debe tener 2 caracteres
    valor = row['VALOR']  # Columna de valor

    # Generar la línea de acuerdo con el formato
    linea = (
        generar_consecutivo(numero_registro + 1) +  # Consecutivo
        "4" +                               # Tipo de registro
        "292" +                             # Número del formato
        "01" +                              # Número de la columna (fijo en 01)
        unidad_captura +                    # Unidad de captura
        subcuenta +                         # Subcuenta
        formatear_valor(valor)              # Valor formateado
    )

    # Agregar la línea a la lista
    if linea not in lineas:  # Evitar duplicados
        lineas.append(linea)

    # Incrementar el consecutivo
    numero_registro += 1

# Agregar el consecutivo final dinámicamente basado en el número de registros
# El consecutivo debe ser el total de registros + 1 y con un "5" al final.
consecutivo_final = generar_consecutivo(numero_registro + 1) + "5"
lineas.append(consecutivo_final)

# Guardar las líneas en un archivo TXT

with open("Capital Minimo COVI.txt", "w") as f:
    for i, linea in enumerate(lineas):
        if i < len(lineas) - 1:
            f.write(linea + "\n")
        else:
            f.write(linea)

